"""
Create Excel file with AI chat agent.
The Excel file contains a button that launches the chat HTML via VBA macro.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

# Model configuration
MODEL = "ServiceNow-AI/Apriel-1.5-15b-Thinker"

# Self-contained HTML chat app
CHAT_HTML = '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>PDFClaw AI Chat</title>
<style>
* { margin: 0; padding: 0; box-sizing: border-box; }
body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #0f0f1a; color: #e0e0e0; height: 100vh; display: flex; flex-direction: column; }
.header { background: #1a1a2e; padding: 12px 16px; border-bottom: 1px solid #2a2a4e; }
.header h1 { font-size: 18px; color: #fff; }
.header p { font-size: 10px; color: #888; margin-top: 2px; }
.api-key-area { background: #16213e; padding: 12px 16px; border-bottom: 1px solid #2a2a4e; display: flex; gap: 8px; align-items: center; }
.api-key-area input { flex: 1; background: #0f0f1a; border: 1px solid #2a2a4e; color: #fff; padding: 8px 12px; border-radius: 4px; font-size: 12px; outline: none; }
.api-key-area input:focus { border-color: #e94560; }
.api-key-area button { background: #e94560; color: #fff; border: none; padding: 8px 16px; border-radius: 4px; font-size: 12px; cursor: pointer; }
.api-key-area button:hover { background: #d63851; }
.api-key-area .saved { color: #7fff7f; font-size: 11px; }
.chat { flex: 1; overflow-y: auto; padding: 12px 16px; display: flex; flex-direction: column; gap: 8px; }
.msg { max-width: 85%; padding: 10px 14px; border-radius: 10px; line-height: 1.4; font-size: 13px; white-space: pre-wrap; word-wrap: break-word; }
.msg.user { background: #e94560; color: #fff; align-self: flex-end; border-bottom-right-radius: 4px; }
.msg.ai { background: #16213e; color: #e0e0e0; align-self: flex-start; border-bottom-left-radius: 4px; }
.msg.system { background: #2a2a4e; color: #888; align-self: center; font-size: 11px; text-align: center; }
.msg.tool { background: #1a3a1a; color: #7fff7f; align-self: flex-start; font-size: 11px; font-family: monospace; }
.msg.error { background: #4a1a1a; color: #ff6b6b; }
.input-area { background: #1a1a2e; padding: 12px 16px; border-top: 1px solid #2a2a4e; display: flex; gap: 8px; }
.input-area input { flex: 1; background: #0f0f1a; border: 1px solid #2a2a4e; color: #fff; padding: 10px 14px; border-radius: 6px; font-size: 13px; outline: none; }
.input-area input:focus { border-color: #e94560; }
.input-area button { background: #e94560; color: #fff; border: none; padding: 10px 20px; border-radius: 6px; font-size: 13px; cursor: pointer; font-weight: 600; }
.input-area button:hover { background: #d63851; }
.input-area button:disabled { background: #555; cursor: not-allowed; }
.typing { color: #888; font-style: italic; }
</style>
</head>
<body>
<div class="header">
    <h1>PDFClaw AI Agent</h1>
    <p>Model: Apriel-1.5-15b-Thinker | Tool Calling | Session Storage</p>
</div>
<div class="api-key-area">
    <input type="password" id="apiKeyInput" placeholder="Enter your Together API key..." />
    <button onclick="saveApiKey()">Save</button>
    <span class="saved" id="apiKeyStatus"></span>
</div>
<div class="chat" id="chat">
    <div class="msg system">Welcome! Enter your Together API key above. Tools: weather, Wikipedia search, fetch URL, calculate, save/load data, time, tweets, generate content.</div>
</div>
<div class="input-area">
    <input type="text" id="input" placeholder="Type your message..." autofocus />
    <button id="sendBtn" onclick="sendMessage()">Send</button>
</div>
<script>
const API_URL = "https://api.together.xyz/v1/chat/completions";
const MODEL = "ServiceNow-AI/Apriel-1.5-15b-Thinker";

const chatEl = document.getElementById("chat");
const inputEl = document.getElementById("input");
const sendBtn = document.getElementById("sendBtn");
const apiKeyInput = document.getElementById("apiKeyInput");
const apiKeyStatus = document.getElementById("apiKeyStatus");

let API_KEY = sessionStorage.getItem('together_api_key') || '';
if (API_KEY) {
    apiKeyInput.value = API_KEY;
    apiKeyStatus.textContent = 'Saved';
}

function saveApiKey() {
    API_KEY = apiKeyInput.value.trim();
    if (API_KEY) {
        sessionStorage.setItem('together_api_key', API_KEY);
        apiKeyStatus.textContent = 'Saved';
    } else {
        apiKeyStatus.textContent = 'Empty';
    }
}

const sessionData = JSON.parse(sessionStorage.getItem('pdfclaw_data') || '{}');

const tools = [
    { type: "function", function: { name: "get_weather", description: "Get current weather for a city", parameters: { type: "object", properties: { city: { type: "string", description: "City name" } }, required: ["city"] } } },
    { type: "function", function: { name: "web_search", description: "Search Wikipedia for information", parameters: { type: "object", properties: { query: { type: "string", description: "Search query" } }, required: ["query"] } } },
    { type: "function", function: { name: "fetch_url", description: "Fetch content from a URL", parameters: { type: "object", properties: { url: { type: "string", description: "URL to fetch" } }, required: ["url"] } } },
    { type: "function", function: { name: "calculate", description: "Perform mathematical calculations", parameters: { type: "object", properties: { expression: { type: "string", description: "Math expression" } }, required: ["expression"] } } },
    { type: "function", function: { name: "save_data", description: "Save data to session storage", parameters: { type: "object", properties: { key: { type: "string" }, value: { type: "string" } }, required: ["key", "value"] } } },
    { type: "function", function: { name: "get_data", description: "Retrieve saved data", parameters: { type: "object", properties: { key: { type: "string" } }, required: ["key"] } } },
    { type: "function", function: { name: "list_data", description: "List all saved data", parameters: { type: "object", properties: {}, required: [] } } },
    { type: "function", function: { name: "get_current_time", description: "Get current date and time", parameters: { type: "object", properties: {}, required: [] } } },
    { type: "function", function: { name: "generate_tweet", description: "Generate a tweet", parameters: { type: "object", properties: { topic: { type: "string" }, style: { type: "string" } }, required: ["topic"] } } },
    { type: "function", function: { name: "generate_content", description: "Generate content", parameters: { type: "object", properties: { type: { type: "string" }, topic: { type: "string" }, length: { type: "string" }, style: { type: "string" } }, required: ["type", "topic"] } } }
];

async function executeTool(toolName, args) {
    switch(toolName) {
        case "get_weather":
            try {
                const geoRes = await fetch('https://geocoding-api.open-meteo.com/v1/search?name=' + encodeURIComponent(args.city) + '&count=1');
                const geoData = await geoRes.json();
                if (!geoData.results || geoData.results.length === 0) return 'City not found';
                const lat = geoData.results[0].latitude;
                const lon = geoData.results[0].longitude;
                const cityName = geoData.results[0].name;
                const weatherRes = await fetch('https://api.open-meteo.com/v1/forecast?latitude=' + lat + '&longitude=' + lon + '&current_weather=true');
                const weatherData = await weatherRes.json();
                const cw = weatherData.current_weather;
                return 'Weather in ' + cityName + ': ' + cw.temperature + 'C, Wind: ' + cw.windspeed + ' km/h';
            } catch(e) { return 'Error: ' + e.message; }
        case "web_search":
            try {
                const res = await fetch('https://en.wikipedia.org/api/rest_v1/page/summary/' + encodeURIComponent(args.query));
                if (!res.ok) return 'Not found on Wikipedia';
                const data = await res.json();
                return 'Wikipedia: ' + data.title + '\\n' + (data.extract || 'No summary');
            } catch(e) { return 'Error: ' + e.message; }
        case "fetch_url":
            try {
                let url = args.url;
                if (!url.startsWith('http')) url = 'https://' + url;
                const res = await fetch('https://api.allorigins.win/raw?url=' + encodeURIComponent(url));
                if (!res.ok) return 'Could not fetch URL';
                const html = await res.text();
                const text = html.replace(/<script[^>]*>[\\s\\S]*?<\\/script>/gi, '').replace(/<style[^>]*>[\\s\\S]*?<\\/style>/gi, '').replace(/<[^>]+>/g, ' ').substring(0, 2000);
                return 'Content: ' + text;
            } catch(e) { return 'Error: ' + e.message; }
        case "calculate":
            try { return 'Result: ' + Function('"use strict"; return (' + args.expression + ')')(); } catch(e) { return 'Error: ' + e.message; }
        case "save_data":
            sessionData[args.key] = args.value;
            sessionStorage.setItem('pdfclaw_data', JSON.stringify(sessionData));
            return 'Saved: ' + args.key;
        case "get_data":
            return sessionData[args.key] ? 'Retrieved: ' + args.key + ' = ' + sessionData[args.key] : 'Not found';
        case "list_data":
            return Object.keys(sessionData).length ? 'Data: ' + JSON.stringify(sessionData) : 'No data saved';
        case "get_current_time":
            return 'Current time: ' + new Date().toLocaleString();
        case "generate_tweet":
            return 'Generate a tweet about "' + args.topic + '" (max 280 chars)';
        case "generate_content":
            return 'Generate a ' + (args.length || 'medium') + ' ' + args.type + ' about "' + args.topic + '"';
        default:
            return 'Unknown tool';
    }
}

let messages = [{ role: "system", content: "You are a helpful AI assistant with access to tools. Use tools when needed. Be concise." }];

inputEl.addEventListener("keydown", function(e) { if (e.key === "Enter" && !sendBtn.disabled) sendMessage(); });

function addMsg(text, cls) {
    const div = document.createElement("div");
    div.className = "msg " + cls;
    div.textContent = text;
    chatEl.appendChild(div);
    chatEl.scrollTop = chatEl.scrollHeight;
    return div;
}

async function sendMessage() {
    const msg = inputEl.value.trim();
    if (!msg) return;
    if (!API_KEY) { addMsg("Please enter your Together API key first!", "system"); apiKeyInput.focus(); return; }
    addMsg(msg, "user");
    inputEl.value = "";
    sendBtn.disabled = true;
    messages.push({ role: "user", content: msg });
    const typing = addMsg("Thinking...", "ai typing");
    try {
        let response = await callAI();
        typing.textContent = response;
        typing.className = "msg ai";
    } catch(e) {
        typing.textContent = "Error: " + e.message;
        typing.className = "msg ai error";
    }
    sendBtn.disabled = false;
    inputEl.focus();
}

async function callAI() {
    const res = await fetch(API_URL, {
        method: "POST",
        headers: { "Content-Type": "application/json", "Authorization": "Bearer " + API_KEY },
        body: JSON.stringify({ model: MODEL, messages: messages, tools: tools, tool_choice: "auto", max_tokens: 1024 })
    });
    if (!res.ok) throw new Error('API Error: ' + res.status);
    const data = await res.json();
    const assistantMsg = data.choices[0].message;
    messages.push(assistantMsg);
    if (assistantMsg.tool_calls && assistantMsg.tool_calls.length > 0) {
        for (const tc of assistantMsg.tool_calls) {
            addMsg('Tool: ' + tc.function.name, "tool");
            const result = await executeTool(tc.function.name, JSON.parse(tc.function.arguments));
            addMsg('Result: ' + result, "tool");
            messages.push({ role: "tool", tool_call_id: tc.id, content: result });
        }
        const finalRes = await fetch(API_URL, {
            method: "POST",
            headers: { "Content-Type": "application/json", "Authorization": "Bearer " + API_KEY },
            body: JSON.stringify({ model: MODEL, messages: messages, max_tokens: 1024 })
        });
        const finalData = await finalRes.json();
        messages.push(finalData.choices[0].message);
        return finalData.choices[0].message.content;
    }
    return assistantMsg.content || "No response";
}
</script>
</body>
</html>'''

def create_excel_file():
    """Create Excel file with instructions"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI Agent"
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 5
    
    # Styles
    title_font = Font(name='Arial', size=24, bold=True, color='FFFFFF')
    subtitle_font = Font(name='Arial', size=12, color='888888')
    header_font = Font(name='Arial', size=14, bold=True, color='E94560')
    normal_font = Font(name='Arial', size=11)
    button_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    
    dark_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
    accent_fill = PatternFill(start_color='E94560', end_color='E94560', fill_type='solid')
    light_fill = PatternFill(start_color='16213E', end_color='16213E', fill_type='solid')
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Title row
    ws.merge_cells('B2:B3')
    ws['B2'] = 'PDFClaw AI Agent for Excel'
    ws['B2'].font = title_font
    ws['B2'].fill = dark_fill
    ws['B2'].alignment = center_align
    
    # Subtitle
    ws['B4'] = 'AI Chat Agent Embedded in Excel Spreadsheet'
    ws['B4'].font = subtitle_font
    ws['B4'].alignment = center_align
    
    # Model info
    ws['B5'] = 'Model: Apriel-1.5-15b-Thinker | Tool Calling Enabled'
    ws['B5'].font = header_font
    ws['B5'].alignment = center_align
    
    # Instructions header
    ws['B7'] = 'How to Use:'
    ws['B7'].font = header_font
    
    # Instructions
    instructions = [
        '1. Enable macros in Excel (see README for details)',
        '2. Click the "Launch AI Chat" button below',
        '3. Your browser will open with the AI chat interface',
        '4. Enter your Together API key (get one free at api.together.xyz)',
        '5. Start chatting with the AI!'
    ]
    
    for i, inst in enumerate(instructions):
        ws[f'B{8+i}'] = inst
        ws[f'B{8+i}'].font = normal_font
    
    # Button area
    ws['B14'] = ''
    ws['B14'].fill = dark_fill
    
    ws.merge_cells('B15:B17')
    ws['B15'] = '🚀 LAUNCH AI CHAT'
    ws['B15'].font = button_font
    ws['B15'].fill = accent_fill
    ws['B15'].alignment = center_align
    
    # Tools section
    ws['B19'] = 'Available Tools:'
    ws['B19'].font = header_font
    
    tools = [
        '• get_weather - Get weather for any city',
        '• web_search - Search Wikipedia for information',
        '• fetch_url - Fetch content from any URL',
        '• calculate - Perform mathematical calculations',
        '• save_data / get_data - Store and retrieve data',
        '• get_current_time - Get current date and time',
        '• generate_tweet - Generate tweet-ready content',
        '• generate_content - Generate articles, emails, code, stories'
    ]
    
    for i, tool in enumerate(tools):
        ws[f'B{20+i}'] = tool
        ws[f'B{20+i}'].font = normal_font
    
    # Example commands
    ws['B29'] = 'Example Commands:'
    ws['B29'].font = header_font
    
    examples = [
        '"What is the weather in Nicosia?"',
        '"Search Wikipedia for Cyprus"',
        '"Calculate 15% of 250"',
        '"Generate an article about AI"'
    ]
    
    for i, ex in enumerate(examples):
        ws[f'B{30+i}'] = ex
        ws[f'B{30+i}'].font = Font(name='Arial', size=11, italic=True)
    
    # Note about macros
    ws['B35'] = 'Note: This file requires VBA macros. See README for setup instructions.'
    ws['B35'].font = Font(name='Arial', size=10, italic=True, color='888888')
    
    # Save
    wb.save('pdfclaw_excel_agent.xlsx')
    print("Excel file created: pdfclaw_excel_agent.xlsx")

def save_vba_module():
    """Save VBA module with embedded HTML"""
    # Escape HTML for VBA
    html_escaped = CHAT_HTML.replace('"', '""').replace('\n', '" & vbCrLf & "')
    
    vba_code = f'''Attribute VB_Name = "AIChatModule"
Option Explicit

Sub LaunchAIChat()
    ' Extract embedded HTML and open in browser
    Dim htmlPath As String
    Dim fso As Object
    Dim file As Object
    Dim htmlContent As String
    
    ' Create temp file
    htmlPath = Environ("TEMP") & "\\pdfclaw_chat.html"
    
    ' HTML content (embedded)
    htmlContent = "{html_escaped}"
    
    ' Write to file
    Set fso = CreateObject("Scripting.FileSystemObject")
    Set file = fso.CreateTextFile(htmlPath, True, True)  ' Unicode=True
    file.Write htmlContent
    file.Close
    
    ' Open in default browser
    CreateObject("Shell.Application").Open htmlPath
End Sub
'''
    
    with open('ai_chat_macro.bas', 'w', encoding='utf-8') as f:
        f.write(vba_code)
    print("VBA module saved: ai_chat_macro.bas")

def save_html_standalone():
    """Save standalone HTML file"""
    with open('chat_app.html', 'w', encoding='utf-8') as f:
        f.write(CHAT_HTML)
    print("HTML chat app saved: chat_app.html")

if __name__ == "__main__":
    create_excel_file()
    save_vba_module()
    save_html_standalone()
    print("\nFiles created:")
    print("  - pdfclaw_excel_agent.xlsx (Excel spreadsheet)")
    print("  - ai_chat_macro.bas (VBA module to import)")
    print("  - chat_app.html (Standalone HTML chat app)")
